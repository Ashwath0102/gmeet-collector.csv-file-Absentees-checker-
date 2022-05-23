from tkinter import *
from tkinter.filedialog import askopenfilename

root = Tk()
root.title("Attendance Absentees Checker")
root.geometry("1100x320")
root.configure(background="SkyBlue4",relief=SUNKEN,bd=16)
Label(root, text="Select the CSV FILE", font=('arial', 25 , 'bold'), background="SkyBlue4", foreground="#FFFFFF").pack()

def browsefunc():
    filename = askopenfilename(filetypes=(("CSV file", "*.csv"),  ("All files", " *.* "),))
    filn=filename.replace('/','\\'+'\\')
    import csv
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    with open(filn) as f:
        reader = csv.reader(f, delimiter=':')
        for row in reader:
            ws.append(row)

    wb.save(filn.replace('.csv','.xlsx'))
    bvc=filn.replace('.csv','.xlsx')
    #print(bvc)
    #import openpyxl
    a = {}
    for i in range(1, 62):
        if i != 31 and i != 50 and i != 24:
            if i < 10:
                a["RA181102602000" + str(i)] = 'U'
            else:
                a["RA18110260200" + str(i)] = 'U'
    #print(a)
    # C:\Users\coola\AppData\Local\Temp\hellow.xlsx
    wb = openpyxl.load_workbook(bvc)
    sheet = wb.active
    current_row = sheet.max_row
    current_column = sheet.max_column
    for j in range(2, current_row + 1):
        if sheet.cell(row=j, column=1).value==None:
            continue
        s = sheet.cell(row=j, column=1).value
        t = 0
        b = 0
        for i in a:
            c = s.find(i)
            if c == -1:
                t = 0
                b = i
            else:
                # print(s)
                t = 1
                b = i
                break
        if t == 0:
            a[b] = 'A'
            # print(s,'A')
            if a[b] == 'U':
                print(s, 'A')
        else:
            a[b] = 'P'
    print("Absentees List")
    lk=[]
    for i in a:
        if a[i] == 'U':
            lk.append(i.replace('RA18110260200',''))
    lk=','.join(lk)
    print(lk)
    global LB
    LB.config(text=str(lk))


#b1=Button(root, text="Browse", font=40, command=browsefunc)
Button(root, text="Browse", bd=5, width=10, height=1, command=browsefunc, font=('arial', 10 * 2, 'bold'), background="SkyBlue4", foreground="#FFFFFF",activebackground="green",activeforeground="snow").pack()
Label(root, text="Absentees Register Numbers", font=('arial', 15 , 'bold'), background="SkyBlue4", foreground="#FFFFFF").pack()
LB=Label(root, text="",relief="groove", bd=5, font=('arial', 10 , 'bold'), background="gray28", foreground="#FFFFFF")
LB.pack()
root.mainloop()